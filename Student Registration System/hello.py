import tkinter
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import openpyxl

root = Tk()
root.geometry("700x500")
root.minsize(600, 500)
root.title("ENTRY FORM")
root.iconbitmap(r'entry.ico')
f1 = Frame(root)
l = Label(f1, text="STUDENT INFORMATION SYSTEM", font="lucida 25 bold", pady=10)
l.pack(pady = 20)
f1.pack()
img = PhotoImage(file="stcet.png")
Label(image=img).pack()
name=StringVar()
add =StringVar()
gname = StringVar()
dept = StringVar()
email = StringVar()
emailadd=StringVar()
day=StringVar()
month=StringVar()
year=StringVar()
phone = StringVar()
ten = StringVar()
twelve = StringVar()

def check(a):
   if a in (',','/','-',' '):
      return True

def check1(a1):
   if a1 in (" "):
      return True

def reset() :
   name.set("")
   add.set("")
   gname.set("")
   dept.set("")
   day.set('1')
   month.set('January')
   year.set('2000')
   email.set('')
   emailadd.set('@gmail.com')
   phone.set("")
   ten.set("")
   twelve.set("")

def newWindow():
   global new
   def back1():
      reset()
      new.destroy()
   new = Toplevel(root)
   new.title("REGISTRATION FORM")
   new.geometry("750x450")
   new.iconbitmap(r'registration.ico')
   name_label = Label(new, text = 'STUDENT NAME ', font=('calibre',12, 'bold'),padx=5,pady=5).grid(row=0,column=0)
   name_entry = Entry(new,textvariable = name, font=('calibre',12,'normal')).grid(row=0,column=1)
   add_label = Label(new, text = 'ADDRESS ', font = ('calibre',12,'bold'),padx=5,pady=5).grid(row=1,column=0)
   add_entry=Entry(new, textvariable = add, font = ('calibre',12,'normal')).grid(row=1,column=1)
   gname_label = Label(new, text = "GUARDIAN'S NAME ", font = ('calibre',12,'bold'),padx=5,pady=5).grid(row=2,column=0)
   gname_entry=Entry(new, textvariable = gname, font = ('calibre',12,'normal')).grid(row=2,column=1)
   dept_label = Label(new, text = 'DEPARTMENT ', font=('calibre',12, 'bold'),padx=5,pady=5).grid(row=3,column=0)
   dept_entry = Entry(new,textvariable = dept, font=('calibre',12,'normal')).grid(row=3,column=1)
   dob_label = Label(new, text = 'DATE OF BIRTH(DD/MM/YYYY)', font=('calibre',12, 'bold'),padx=5,pady=5).grid(row=4,column=0)
   list1=[]
   for i in range(0,31):
      list1.append(i+1)
   combobox=ttk.Combobox(new,textvariable=day,values=list1)
   combobox.set('1')
   combobox.grid(row=4, column=1)
   list2=['January','February','March','April','May','June','July','August','September','October','November','December']
   combobox=ttk.Combobox(new,textvariable=month,values=list2)
   combobox.set('January')
   combobox.grid(row=4, column=2)
   list3=[]
   for i in range(2000,2008):
      list3.append(i+1)
   combobox=ttk.Combobox(new,textvariable=year,values=list3)
   combobox.set('2000')
   combobox.grid(row=4, column=3,pady=10,padx=10)
   email_label = Label(new, text = 'EMAIL ID ', font=('calibre',12, 'bold'),padx=5,pady=5).grid(row=5,column=0)
   email_entry = Entry(new,textvariable = email, font=('calibre',12,'normal')).grid(row=5,column=1)
   list4=['@hotmail.com','@gmail.com','@yahoo.com']
   combobox=ttk.Combobox(new,textvariable=emailadd,values=list4)
   combobox.set('@gmail.com')
   combobox.grid(row=5, column=2,padx=10,pady=10)
   phone_label = Label(new, text = 'MOBILE NUMBER  (+91)', font=('calibre',12, 'bold'),padx=5,pady=5).grid(row=6,column=0)
   phone_entry = Entry(new,textvariable = phone, font=('calibre',12,'normal')).grid(row=6,column=1)
   ten_label = Label(new, text = 'CLASS 10 PERCENTAGE ', font=('calibre',12, 'bold'),padx=5,pady=5).grid(row=7,column=0)
   ten_entry = Entry(new,textvariable = ten, font=('calibre',12,'normal')).grid(row=7,column=1)
   twelve_label = Label(new, text = 'CLASS 12 PERCENTAGE ', font=('calibre',12, 'bold'),padx=5,pady=5).grid(row=8,column=0)
   twelve_entry = Entry(new,textvariable = twelve, font=('calibre',12,'normal')).grid(row=8,column=1)

   bb1 = Button(new,text="SUBMIT",font=('arial', 12, 'bold'),command=enter_data).grid(row=10,column=0,pady=10)
   bb2 = Button(new,text="RESET",font=('arial', 12, 'bold'),command=reset).grid(row=10,column=1,pady=10)
   bb3 = Button(new,text="BACK",font=('arial', 12, 'bold'),command=back1).grid(row=10,column=2,pady=10)

def enter_data():
   new.withdraw()
   result=messagebox.askquestion("Submit","Do you want to submit")
   new.deiconify()
   if result=='yes':
      nam=name.get()
      address=add.get()
      gnam=gname.get()
      depmt=dept.get()
      Day=day.get()
      Month=month.get()
      Year=year.get()
      eml=email.get()
      dob=Day+"/"+Month+"/"+Year
      Emailadd=emailadd.get()
      emailaddress=eml+Emailadd
      ph=phone.get()
      fullph=ph
      tenn=ten.get()
      twelvee=twelve.get()
      if name and address and gname and dept and dob and email and ph and ten and twelve:
         if (len(fullph)!=10) or fullph.isdigit()==False:
            new.withdraw()
            phone.set("")
            messagebox.showerror(title="Error",message="Phone number is incorrect!!!")
            new.deiconify()
         elif nam.isalpha()==False and check1(nam):
            new.withdraw()
            name.set("")
            messagebox.showerror(title="Error",message="Name should be in characters")
            new.deiconify()
         elif address.isalnum()==False and check(address):
            new.withdraw()
            add.set("")
            messagebox.showerror(title="Error",message="Enter a valid address")
            new.deiconify()
         elif gnam.isalpha()==False and check1(gnam):
            new.withdraw()
            gname.set("")
            messagebox.showerror(title="Error",message="Guardian name should be in characters")
            new.deiconify()
         elif depmt.isalpha()==False and check1(depmt):
            new.withdraw()
            dept.set("")
            messagebox.showerror(title="Error",message="Department should be in characters")
            new.deiconify()
         elif ' ' in eml :
            messagebox.showerror(title="error",message="Enter a valid email address")
            new.withdraw()
            email.set("")
            new.deiconify()
         elif tenn.isdigit()==False or twelvee.isdigit()==False:
            new.withdraw()
            ten.set("")
            twelve.set("")
            messagebox.showerror(title="Error",message="Ten and Twelve's percentage should be in digits")
            new.deiconify()
         elif int(tenn)<0 or int(tenn)>100:
            new.withdraw()
            ten.set("")
            messagebox.showerror(title="Error",message="Ten's percentage should be within 0 to 100")
            new.deiconify()
         elif int(twelvee)<0 or int(twelvee)>100:
            new.withdraw()
            twelve.set("")
            messagebox.showerror(title="Error",message="Twelve's percentage should be within 0 to 100")
            new.deiconify()
         else:
            wb=openpyxl.load_workbook("student.xlsx")
            sheet=wb.active
            r=sheet.max_row
            r_id='23/'+str(r)
            sheet.append([nam.upper(),address.upper(),gnam.upper(),depmt.upper(),dob,emailaddress,fullph,tenn,twelvee,r_id])
            wb.save("student.xlsx")
            new.withdraw()
            messagebox.showinfo(title="Message",message="Your form is successfully submitted and your registration id is"+r_id)
            reset()
            b1['state']=DISABLED
      else:
         new.withdraw()
         reset()
         messagebox.showerror(title="Error",message="Please fill all data!!!")
         new.deiconify()

def display():
   Nw1=Toplevel(root)
   Nw1.title("DISPLAYING INFORMATION")
   Nw1.geometry("750x650")
   Nw1.iconbitmap(r'display.ico')
   wb=openpyxl.load_workbook("student.xlsx")
   sheet=wb.active
   list_values=list(sheet.values)
   cols=list_values[0]
   tree=ttk.Treeview(Nw1,column=cols,show="headings")
   treescrolly = tkinter.Scrollbar(Nw1, orient="vertical", command=tree.yview)
   treescrollx = tkinter.Scrollbar(Nw1, orient="horizontal", command=tree.xview)
   tree.configure(xscrollcommand=treescrollx.set, yscrollcommand=treescrolly.set)
   treescrollx.pack(side="bottom", fill="x")
   treescrolly.pack(side="right", fill="y")

   for col_name in cols:
      tree.heading(col_name,text=col_name)
   tree.pack(expand=TRUE,fill='y')

   for value_tuple in list_values[1:]:
      tree.insert('',tkinter.END,values=value_tuple)

f = Frame(root, pady=10)
b1=Button(f, text="ADD", font="arial 20 bold", padx=45, borderwidth=5, relief=GROOVE, command=newWindow)
b1.grid(row=15, column=1)
b2=Button(f, text="DISPLAY", font="arial 20 bold", padx=45, borderwidth=5, relief=GROOVE, command=display)
b2.grid(row=15, column=3)
b3=Button(f, text="EXIT", font="arial 20 bold", padx=45, borderwidth=5, relief=GROOVE, command=root.destroy)
b3.grid(row=15, column=5)
f.pack(side=BOTTOM)

root.mainloop()